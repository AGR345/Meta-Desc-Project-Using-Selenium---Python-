import openpyxl
import time

start_time = time.time()


excel_source_file_path = "Meta-desc-900-pages_final_updated.xlsx"
wb_obj = openpyxl.load_workbook(excel_source_file_path)
sheet_obj = wb_obj.active

max_col = sheet_obj.max_column
max_row = sheet_obj.max_row

print('***********************************')
print('***********************************')
print('***********************************')
print(f'No of Actual Columns in the excelsheet are: {max_col}')
print(f'No of Actual Rows in the excelsheet are: {max_row}')
# 6 is exclude top and bottom rows in iteration
max_row_tobe_considered = int(max_row) - 6
# print(f'Actual Rows are: {max_row}')
print(f'Rows To Be Considered are:   {max_row_tobe_considered-1}')
print('***********************************')
print('***********************************')
print('***********************************')
print('############################################################')
row_start_num = input('Type Row Number from where """NAKSHATRA""" should run (From "Begining > give 1" else, "row_no-1") ? : ')
print('############################################################')
# implicit_wait_seconds_dev = 45
# implicit_wait_seconds_live = 15
implicit_wait_seconds_dev = 90
implicit_wait_seconds_live = 45

for i in range(int(row_start_num)+1, max_row_tobe_considered+1):
    import time
    each_record_start_time = time.time()
    URL = sheet_obj.cell(row=i, column=4)
    URL = URL.value
    meta_desc_value = sheet_obj.cell(row=i, column=7)
    meta_desc_value = meta_desc_value.value
    try:
        meta_desc_value = meta_desc_value.strip()
    except:
        meta_desc_value = meta_desc_value
        print('Excel Meta Desc value is blank')
        auto_comments = sheet_obj.cell(row=i, column=13)
        auto_comments.value = 'Excel Meta Desc value is blank'
        wb_obj.save(excel_source_file_path)

    ################################################################################
    from selenium import webdriver  # To activate the webdriver
    # To maximize the chrome browser window
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.keys import Keys  # keys to send
    from selenium.webdriver import ActionChains   # rightclick actions
    import time  # For using sleep()
    from selenium.webdriver.support.select import Select

    chrome_options = Options()
    # for Chrome Maximized size
    chrome_options.add_argument('--start-maximized')
    # Disabling 'Chrome is being controlled by automated test software' popup on chrome
    chrome_options.add_argument("--disable-infobars")

    import openpyxl

    browser = webdriver.Chrome('\\chromedriver.exe', chrome_options=chrome_options)

    ###############################################################################
    # DEV MODE FOR POSTING
    ###############################################################################
    try:
        # browser opend with DEV URL
        browser.implicitly_wait(int(implicit_wait_seconds_dev))
        browser.get(URL)
        time.sleep(3)

        print('*******************************************************')
        print('*******************************************************')
        print('*******************************************************')
        print(f'({i-1}) Opened URL:  {URL}')
        print('*******************************************************')

        # FINDING PRODUCT NAME IN FOR EACH REORD - INITIALIZATION
        time.sleep(1)
        name_from_Dev_posting = browser.find_element_by_css_selector('#Name')
        name_from_Dev_posting = name_from_Dev_posting.get_attribute('value')
        print(f'Proudct Name:  {name_from_Dev_posting}')

        # FINDING METADESCRIPTION BOX TO UPDATE NEW META DESC VALUE
        time.sleep(1)
        meta_desc_box_Fld = browser.find_element_by_css_selector(
            '#MetaDescription')
        time.sleep(1)
        meta_desc_box_Fld.clear()
        print('Cleared Meta Desc Field')
        meta_desc_box_Fld.send_keys(meta_desc_value)
        time.sleep(1)
        print('Updated Meta Desc Field')

        # SAVING THE UDPATED META DESC VALUE USING SAVE CHANGES BUTTON
        save_button = browser.find_element_by_css_selector(
            '#form0 > div > div.panel-footer > div > div.col-md-offset-1.col-md-9 > button.btn.btn-primary')
        save_button.click()
        print('Clicked Save Button')
        DEV_AUTO_STATUS = sheet_obj.cell(row=i, column=10)
        DEV_AUTO_STATUS.value = 'COMPLETED'
        wb_obj.save(excel_source_file_path)
        print('***********************************')

        ##################################################################
        # PUBLISHNG CODE STARTS HERE
        ##################################################################
        try:
            # CLICKING PUBLISH BUTTON ON THE TOP RIGHT CORNER
            time.sleep(1)
            publish_button = browser.find_element_by_css_selector(
                '#wrapper > nav > div > div.collapse.navbar-collapse > ul.nav.navbar-nav.navbar-right > li:nth-child(3) > a')
            publish_button.click()
            print('Entered to PUB from DEV')
            time.sleep(1)

            # CLICKING ON ADVANCED SEARCH BUTTON FOR FILETEING PURPOSE
            advanced_search_button = browser.find_element_by_css_selector(
                '#changes_wrapper > div:nth-child(1) > div.col-md-7.text-center > div > a.btn.btn-default.btn-advanced-search > span')
            advanced_search_button.click()
            print('Clicked on "Advance Search" button in PUB MODE')
            time.sleep(2)

            # SELECTING 'PRODUCTS' FROM THE DROP DWON MENU
            product_dropdown_button = Select(browser.find_element_by_css_selector(
                '#changes_wrapper > div.advanced-search.collapse.in > div > div > select:nth-child(4)'))
            product_dropdown_button.select_by_visible_text('Products')
            print('Selected "Products" from the drop down menu')
            time.sleep(1)

            # FINDING FOR SEARCH BOX FOR CORRECT PRODUCT NAME UNDER PRODUCTS FILTER ON PUBLISH MODE PAGE
            search_field_button = browser.find_element_by_css_selector(
                '#changes_filter > label > input')
            time.sleep(1)
            search_field_button.clear()
            print('Cleared Search Field')
            search_field_button.send_keys(name_from_Dev_posting)
            time.sleep(1)

            # CHECKING FOR 'NO MATCHING RECORDS FOUND' OR 'EXACT RECORD' SELECTION USING 'CHECK BOX' OPTION
            matches_text = browser.find_element_by_css_selector(
                '#changes > tbody > tr > td')
            error_no_pages_to_publish = matches_text.text
            print('################################')
            # if error_no_pages_to_publish != None:
            #     print(f'Avaiable Records Status for PUBLISH: {error_no_pages_to_publish}')
            # print('################################')
            time.sleep(1)
            if (error_no_pages_to_publish == 'No matching records found') or ('no record' in error_no_pages_to_publish) or ('record' in error_no_pages_to_publish):
                url_id = URL.split('/')
                # print(url_id)
                print('********************************************')
                print(
                    f'fOR THIS PRODUCT ID : {url_id[6]} - NO NEW META DESC DATA UPDATED (PUBLISH NOT REQUIRED)')
                print('********************************************')
                PUBLISH_AUTO_STATUS = sheet_obj.cell(row=i, column=11)
                PUBLISH_AUTO_STATUS.value = 'NO NEW CONTENT UPDATED - PUBLISH NOT REQUIRED'
                wb_obj.save(excel_source_file_path)
                browser.close()
            else:

                url_id = URL.split('/')
                # SELECTING A CHECK BOX FOR ITS APPROPRIATE PUB NAME OR PRODUCT ID
                product_check_button = browser.find_element_by_css_selector(
                    '#changes > tbody > tr:nth-child(1) > td.select-checkbox.selection-cell')
                product_check_button.click()
                print(
                    f'Selected {url_id[6]} GMN product and its name is : {name_from_Dev_posting} ')
                time.sleep(2)

                # CLICKING ON 'PUBLISH SELECTED CHAGNES' BUTTON ON THE LEFT CORNER - IIIrd ROW
                publish_selected_changes_button = browser.find_element_by_css_selector(
                    '#confirm-publish')
                publish_selected_changes_button.click()
                print(f'"Published Changes Selected" Button Clicked ')
                time.sleep(2)

                # CLICKING ON CONFIRM PUBLISH BUTTON - MODAL (DIALOG BOX)
                time.sleep(2)
                confirm_Publish_button = browser.find_element_by_css_selector(
                    '#confirm-publish-modal > div > div > div.modal-footer > button.btn.btn-primary')
                confirm_Publish_button.click()
                #confirm-publish-modal > div > div > div.modal-footer > button.btn.btn-primary

                # FINAL STAAGE OF PUBLISH MODE TO PUT PUBLISH STATUS TO EXCEL
                url_id = URL.split('/')
                # print(url_id)
                print(f'SUCCESSFULLY PUBLISHED - {url_id[6]} GMN PRODUCT PAGE')
                PUBLISH_AUTO_STATUS = sheet_obj.cell(row=i, column=11)
                PUBLISH_AUTO_STATUS.value = 'COMPLETED'
                wb_obj.save(excel_source_file_path)
                browser.close()

        except:
            print(
                f'Unable to Publish {url_id[6]} GMN product and its name is : {name_from_Dev_posting} ')
            print(
                f' Loop broken while publishing at row no: \n{(i - 1)}: {URL}  \n {meta_desc_value}\n')
            PUBLISH_AUTO_STATUS = sheet_obj.cell(row=i, column=11)
            PUBLISH_AUTO_STATUS.value = 'ERROR - WHILE PUBLISHING'
            wb_obj.save(excel_source_file_path)
            browser.close()

        try:

            url_id = URL.split('/')
            GMN_NUMBER = url_id[6]
            # print(url_id[6])
            # print(f'view-source:https://www.eastman.com/Pages/ProductHome.aspx?product={url_id[6]}')

            from selenium import webdriver  # To activate the webdriver
            # To maximize the chrome browser window
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.common.keys import Keys  # keys to send
            from selenium.webdriver import ActionChains   # rightclick actions
            import time  # For using sleep()
            from selenium.webdriver.support.select import Select
            from bs4 import BeautifulSoup
            import requests

            chrome_options = Options()
            # for Chrome Maximized size
            chrome_options.add_argument('--start-maximized')
            # Disabling 'Chrome is being controlled by automated test software' popup on chrome
            chrome_options.add_argument("--disable-infobars")

            browser = webdriver.Chrome('\\chromedriver.exe', chrome_options=chrome_options)
            live_url = f'https://www.eastman.com/Pages/ProductHome.aspx?product={GMN_NUMBER}'
            live_url = str(live_url)

            browser.implicitly_wait(int(implicit_wait_seconds_live))
            live_source_code = browser.get(live_url)

            live_meta_desc_html_source = browser.page_source
            # print(live_meta_desc_html_source)
            time.sleep(2)
            # using fetched selenium 'view_source_code' via beautifulsoup scraping
            html_soup = BeautifulSoup(live_meta_desc_html_source, 'lxml')

            for meta_desc_value_udpatedOnLive in html_soup.find_all('meta'):
                try:
                    meta_desc_value_udpatedOnLive = meta_desc_value_udpatedOnLive.get(
                        'content')
                    # print(meta_desc_value_udpatedOnLive)
                except:
                    print('META DESCRIPTION - CONTENT IS NOT AVAILABLE ON LIVE PAGE')

            ################################################################################
            ################################################################################

            print('#######################################################')
            print(f'DEV URL is:   {URL}')
            print(f'This meta desc value is from excel \n{meta_desc_value}')
            print('#######################################################')
            print(f'Live URL is:   {live_url}')
            print(
                f'This meta desc value is from Live page \n{meta_desc_value_udpatedOnLive}')
            print('#######################################################')
            time.sleep(1)
            ################################################################################
            ################################################################################
            if meta_desc_value == meta_desc_value_udpatedOnLive:
                url_id = URL.split('/')
                # print('')
                print(
                    '###################################################################################')
                print(
                    f'{url_id[6]} GMN PRODUCT IS VERIFIED ONLINE VS EXCEL META DESC')
                print(
                    '###################################################################################')
                print(f'Record {i-1} is closed.')
                # print()
                # print()
                # print()
                live_AUTO_STATUS = sheet_obj.cell(row=i, column=12)
                live_AUTO_STATUS.value = 'COMPLETED'
                wb_obj.save(excel_source_file_path)
                browser.close()
            else:
                url_id = URL.split('/')
                # print('\n\n')
                print(
                    '###################################################################################')
                print(
                    f'{url_id[6]} GMN PRODUCT LIVE META DESC IS NOT MATCHING WITH VS EXCEL META DESC')
                print(
                    '###################################################################################')
                print(f'Record {i-1} is closed.')
                # print()
                # print()
                # print()
                live_AUTO_STATUS = sheet_obj.cell(row=i, column=12)
                live_AUTO_STATUS.value = 'EXCEL META VS LIVE META - NO MATCH'
                wb_obj.save(excel_source_file_path)
                browser.close()

        except:
            url_id = URL.split('/')
            print(f'Unable to open the Live URL for:  {GMN_NUMBER}')
            time.sleep(3)
            ############################################
            live_AUTO_STATUS = sheet_obj.cell(row=i, column=12)
            live_AUTO_STATUS.value = 'TRY AGAIN'
            wb_obj.save(excel_source_file_path)
            # browser.close()
            ############################################
            browser.close()

    except:
        print(f' Loop broken at row no: {(i - 1)}: {URL}  | {meta_desc_value}')
        DEV_AUTO_STATUS = sheet_obj.cell(row=i, column=10)
        DEV_AUTO_STATUS.value = 'TRY AGAIN - 404/ERROR'
        wb_obj.save(excel_source_file_path)
        browser.close()

    #################################################################################################
    # FINAL STATUS FOR EXCELSHEET PURPOSE TO FIND OUT HOW MANY COMPLETED AND INCOMPLETE
    #################################################################################################
    final_requests_status_value = sheet_obj.cell(row=i, column=14)
    if ((DEV_AUTO_STATUS.value == 'COMPLETED') and (PUBLISH_AUTO_STATUS.value == 'NO NEW CONTENT UPDATED - PUBLISH NOT REQUIRED') and (live_AUTO_STATUS.value == 'COMPLETED')):
        # final_requests_status_value == sheet_obj.cell(row=i, column=14)
        final_requests_status_value.value = 'COMPLETED'
        wb_obj.save(excel_source_file_path)
        browser.quit()
    elif ((DEV_AUTO_STATUS.value == 'COMPLETED') and (PUBLISH_AUTO_STATUS.value == 'COMPLETED') and (live_AUTO_STATUS.value == 'COMPLETED')):
        # final_requests_status_value == sheet_obj.cell(row=i, column=14)
        final_requests_status_value.value = 'COMPLETED'
        wb_obj.save(excel_source_file_path)
        browser.quit()
    else:
        # final_requests_status_value == sheet_obj.cell(row=i, column=14)
        final_requests_status_value.value = 'TRY AGAIN'
        wb_obj.save(excel_source_file_path)
        browser.quit()

    # COUNTING TIME FOR EACH RECORD
    each_record_complete_end_time = time.time()
    complete_elapsed_time = each_record_complete_end_time - each_record_start_time
    # complete_elapsed_time = f'{round(complete_elapsed_time)} Seconds'
    # complete_elapsed_time = round(complete_elapsed_time)
    print('###################################################################################')
    print('###################################################################################')
    print(
        f'Total time taken to complete this record is:  {complete_elapsed_time}')
    print('###################################################################################')
    print('###################################################################################')

    each_record_Trans_Completion_time_value = sheet_obj.cell(row=i, column=15)
    each_record_Trans_Completion_time_value.value = complete_elapsed_time
    wb_obj.save(excel_source_file_path)
    # browser.quit()

    continue

# COUNTING TIME FOR COMPLETE REQUEST
# complete_end_time = time.time()
# complete_elapsed_time = complete_end_time - start_time
# print(f'Script started at  :  {start_time}')
# print(f'Complete Script ended at :  {complete_end_time}')
# print(f'Total time taken to complete this task is:  {complete_elapsed_time}')
# print('###################################################################################')
# print('###################################################################################')
# print()
# print()
# print()
